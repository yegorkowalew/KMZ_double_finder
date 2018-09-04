from datetime import datetime
import time
import os
import inspect
from copy import copy
# from modules.dirs.dir_work import user_input_work
import openpyxl

def work_with_dir():
    folder_for_find_xlsx = os.path.dirname(os.path.abspath(inspect.stack()[0][1])) # полный путь к папке из которой выполняется файл
    folder_for_find_xlsx = folder_for_find_xlsx+"\\files"
    list_files = [] # список файлов с полными путями
    for file in os.listdir(folder_for_find_xlsx): # во всех файлах папки
            if file.endswith(".xlsx"): # находим файл совпадающий с шаблоном
                list_files.append(os.path.join(folder_for_find_xlsx, file)) # добавляем файл, с полным путем, подходящий шаблону, в список
    return list_files

def file_to_wb(file_name):
    return openpyxl.load_workbook(filename = file_name)

double_sheet_name = "Сортировка по узлам"

if __name__ == '__main__':
    start_time = time.process_time() # Засекаем время
    xls_file = work_with_dir()[0] # первый файл в найденых файлах

    wb = openpyxl.Workbook()
    new_sheet = wb.create_sheet('сумма') # создаем новый лист в файле

    i=0
    for need_file in work_with_dir():
        print('Работаю с файлом: ' + need_file)
        work_wb = file_to_wb(need_file)
        sheet = work_wb[double_sheet_name]
        
        for row in range(6, sheet.max_row):
            i+=1
            new_sheet.cell(row=i, column=1).value = sheet.cell(row=row, column=2).value
            new_sheet.cell(row=i, column=2).value = sheet.cell(row=row, column=3).value
            new_sheet.cell(row=i, column=3).value = sheet.cell(row=row, column=6).value
            new_sheet.cell(row=i, column=4).value = sheet.cell(row=row, column=9).value
            new_sheet.cell(row=i, column=5).value = sheet.cell(row=row, column=10).value
            # new_sheet.cell(row=i, column=6).value = sheet.cell(row=row, column=7).value
            # new_sheet.cell(row=i, column=7).value = sheet.cell(row=row, column=8).value
            # new_sheet.cell(row=i, column=8).value = sheet.cell(row=row, column=3).value
            # for coll in range(1, 11):
                # print(sheet.cell(row=row, column=coll).value)
                # sheet.cell(row=row, column=coll).value
                # new_sheet.cell(row=i, column=coll)._style = copy(sheet.cell(row=row, column=coll)._style)
                


    wb.save(os.path.dirname(os.path.abspath(inspect.stack()[0][1]))+'\\file.xlsx')