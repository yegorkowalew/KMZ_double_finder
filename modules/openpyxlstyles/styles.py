from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

# Заливка серым сплошным цветом
fill_1 = PatternFill(fgColor='D9D9D9', fill_type='solid')

# Заливка желтым цветом
fill_2 = PatternFill(fgColor='F7E2AE', fill_type='solid')

# Обрамление черными, тонкими границами
border_1 = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

# Основной шрифт
font_1 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

# Повернуть шрифт вертикально
alignment_1 = Alignment(text_rotation=180,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0
                        )
