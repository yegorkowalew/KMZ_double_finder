from modules.selflog.selflog import logger
from openpyxl.comments import Comment

import openpyxl
from copy import copy
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from modules.openpyxlstyles.styles import fill_1, fill_2, border_1, font_1, alignment_1

from datetime import datetime
import time
import os
import inspect

# pyinstaller find_doubles_2.py --onefile
def big_trabl(xls_file, trabl):
    logger.critical('Что-то пошло не так при первичной обработке файла: %s. \nПитончег говорит: %s' %
          (xls_file, trabl))
    time.sleep(60)
    exit(0)

exit_sum_file = 'Итоговый файл.xlsx'
exit_files_dir = 'Просумированные'
first_row = 6
end_coll = 0
second_row = 0
after_column = 8
double_sheet_name = "Сортировка по узлам"
shop_list = []

end_names_original = {
    'npp': [1, "№ п/п", 3.7, ],
    'types': [2, "Наименование", "21", ],
    'name': [3, "Позиция", "36", ],
    'knot': [4, "Кол-во узел (шт.)", 3.7, ],
    'product': [5, "Кол-во изделие (шт.)", 3.7, ],
    'order': [6, "Кол-во заказ (шт.)", 3.7, ],
    'nzp': [7, "Кол-во НЗП (шт.)", 3.7, ],
    'massiv': [8, "Массив", 3.7, ],
    'size_clean': [9, "Размеры чист.", "26", ],
    'size_workpiece': [10, "Размеры заготовки", "26", ],
    'operations': [11, "Операции", "26", ],
    'shop': [12, "Розцеховка", "12", ],
}
end_names = end_names_original


def if_num(cell):
    try:
        eval(cell)
        return True
    except:
        return False

class Unit:
    def __init__(self, types, name, knot, product, order, nzp, massiv, size_clean, size_workpiece, operations, shop, row):
        """Constructor for Unit"""
        self.types = types  # Столбец "Наименование" (не обязательный)
        self.name = name  # Столбец "Позиция" (обязательный)
        self.knot = knot  # Столбец "Узел" (не обязательный)
        self.product = product  # Столбец "Изделие" (не обязательный)
        self.order = order  # Столбец "Заказ" (обязательный)
        self.nzp = nzp  # Столбец: "Кол-во НЗП" (указывать не обязательно)
        self.massiv = massiv  # Столбец: "Массив" (указывать не обязательно)
        # Столбец: "Размеры чист." Описание узла (указывать не обязательно, но желательно)
        self.size_clean = size_clean
        # Столбец: "Размеры заготовки" (указывать не обязательно)
        self.size_workpiece = size_workpiece
        # Столбец: "Операции" (указывать не обязательно)
        self.operations = operations
        self.shop = shop  # Столбец: "Розцеховка" (указывать не обязательно)
        self.row = row  # поле для хранения номера строки

class Detail:
    def __init__(self, npp, types, name, knot, product, order, nzp, massiv, size_clean, size_workpiece, operations, shop, doubles, row, unit_row):
        """Constructor for Detail"""
        self.npp = npp  # "п/п" - Номер по порядку детали (Указывать не обязательно, желательно самому пронумеровать)
        # Столбец: "Наименование" тип узла (указывать обязательно)
        self.types = types
        self.name = name  # Столбец "Позиция" (обязательный)
        self.knot = knot  # Столбец "Узел" (не обязательный)
        self.product = product  # Столбец "Изделие" (не обязательный)
        if if_num(order):
            # Столбец: "Кол-во заказ" по этому полю будут подсчитываться суммы дубликатов (указывать обязательно)
            self.order = eval(order)
        else:
            self.order = int(order)
        self.nzp = nzp  # Столбец: "Кол-во НЗП" (указывать не обязательно)
        self.massiv = massiv  # Столбец: "Массив" (указывать не обязательно)
        # Столбец: "Размеры чист." (указывать не обязательно)
        self.size_clean = size_clean
        # Столбец: "Размеры заготовки" (указывать не обязательно)
        self.size_workpiece = size_workpiece
        # Столбец: "Операции" (указывать не обязательно)
        self.operations = operations
        # self.shop Столбец: "Розцеховка" таблица будет разбита на листы в которых каждый цех по отдельности (указывать обязательно)
        # разбиваем строку вида 104-101, на список [104, 101]
        self.shop = str(shop).split("-")
        for st in self.shop:
            st = st.replace('/', '')
        # print(self.shop)
        self.doubles = doubles  # список дубликатов
        self.row = row  # поле для хранения номера строки
        self.unit_row = unit_row  # Указываем строку родительского узла

def work_with_dir(folder_for_find_xlsx):
    # folder_for_find_xlsx = os.path.dirname(os.path.abspath(inspect.stack()[0][1])) # полный путь к папке из которой выполняется файл
    list_files = []  # список файлов с полными путями
    for file in os.listdir(folder_for_find_xlsx):  # во всех файлах папки
        if file.endswith(".xlsx"):  # находим файл совпадающий с шаблоном
            # добавляем файл, с полным путем, подходящий шаблону, в список
            list_files.append(os.path.join(file))
    if list_files == []:
        logger.critical('Нету файлов для обработки')
        time.sleep(30)
        exit(0)
    else:
        if exit_sum_file in list_files:
            list_files.remove(exit_sum_file)
            logger.warning('Файл "%s" уже есть в папке, буду перезаписывать' % (exit_sum_file))
        for_del = []
        for list_file in list_files:
            if list_file[0] == '~':
                logger.warning('С файлом "%s" работать не буду' % (list_file))
                for_del.append(list_file)
        for for_del_file in for_del:
            list_files.remove(for_del_file)
    return list_files

def file_to_wb(file_name):
    try:
        return openpyxl.load_workbook(filename=file_name)
    except:
        logger.critical('Ошибка открытия файла: %s' % (file_name))
        # time.sleep(30)
        exit(0)

def work(work_filename, work_dir, exit_files_dir):
    new_sheet_list = []
    start_time = time.process_time()  # Засекаем время
    work_wb = file_to_wb(work_filename)  # из файла открываем книгу
    sheet = work_wb[work_wb.sheetnames[0]]  # берем первый лист в файле
    if double_sheet_name in work_wb.sheetnames:
        double_sheet = work_wb[double_sheet_name]
        work_wb.remove(double_sheet)
        double_sheet = work_wb.create_sheet(
            double_sheet_name)  # создаем новый лист в файле
        new_sheet_list.append(double_sheet)
    else:
        double_sheet = work_wb.create_sheet(
            double_sheet_name)  # создаем новый лист в файле
        new_sheet_list.append(double_sheet)

    units = []
    details = {}

    for row in range(1, sheet.max_row):  # нашел первую строку в таблице
        if sheet.cell(row=row, column=end_names.get('npp')[0]).value == "Уз.":
            first_row = row
            break

    for row in range(first_row, sheet.max_row):  # нашел последнюю строку в таблице
        if sheet.cell(row=row, column=end_names.get('name')[0]).value:
            second_row = row+1

    logger.info('Последняя обрабатываемая строка в таблице:  %s' % (str(second_row)))

    # end_coll = 0
    for col in range(1, sheet.max_column): # нашел последний столбец в таблице
        end_coll = col

    for row in range(first_row, second_row+1):
        if sheet.cell(row=row, column=1).value == "Уз.":
            logger.info('Нашел узел: %s' % (str(sheet.cell(row=row, column=3).value)))
            unit_row = row
            npp = 0
            unit = Unit(
                sheet.cell(row=row, column=end_names.get('types')[0]).value,
                sheet.cell(row=row, column=end_names.get('name')[0]).value,
                sheet.cell(row=row, column=end_names.get('knot')[0]).value,
                sheet.cell(row=row, column=end_names.get('product')[0]).value,
                sheet.cell(row=row, column=end_names.get('order')[0]).value,
                sheet.cell(row=row, column=end_names.get('nzp')[0]).value,
                sheet.cell(row=row, column=end_names.get('massiv')[0]).value,
                sheet.cell(row=row, column=end_names.get(
                    'size_clean')[0]).value,
                sheet.cell(row=row, column=end_names.get(
                    'size_workpiece')[0]).value,
                sheet.cell(row=row, column=end_names.get(
                    'operations')[0]).value,
                sheet.cell(row=row, column=end_names.get('shop')[0]).value,
                row,)
            units.append(unit)
        else:
            if sheet.cell(row=row, column=end_names.get('name')[0]).value:
                logger.info('Нашел деталь: %s' % (str(sheet.cell(row=row, column=3).value)))
                npp += 1
                unit = Detail(
                    npp,
                    sheet.cell(row=row, column=end_names.get('types')[0]).value,
                    sheet.cell(row=row, column=end_names.get('name')[0]).value,
                    sheet.cell(row=row, column=end_names.get('knot')[0]).value,
                    sheet.cell(row=row, column=end_names.get('product')[0]).value,
                    sheet.cell(row=row, column=end_names.get('order')[0]).value,
                    sheet.cell(row=row, column=end_names.get('nzp')[0]).value,
                    sheet.cell(row=row, column=end_names.get('massiv')[0]).value,
                    sheet.cell(row=row, column=end_names.get(
                        'size_clean')[0]).value,
                    sheet.cell(row=row, column=end_names.get(
                        'size_workpiece')[0]).value,
                    sheet.cell(row=row, column=end_names.get(
                        'operations')[0]).value,
                    sheet.cell(row=row, column=end_names.get('shop')[0]).value,
                    [],
                    row,
                    unit_row)
                details.update({row: unit})

    for detail_for_add in details.values():  # нашел дубликаты и дописал их в поле .doubles
        for detail in details.values():
            if detail_for_add.name == detail.name:
                detail_for_add.doubles.append(detail.row)

    after_end_names = {}
    m_colum = 0
    # Новый массив after_end_names со вставлеными столбцами уз.
    for key, value in end_names.items():
        if value[0] < after_column+1:
            after_end_names.update({key: value})
        else:
            if value[0] == after_column+1:
                m_colum = value[0]
                for i in units:
                    after_end_names.update({i.row: [m_colum, i.name, 3.7, ]})
                    m_colum += 1
            if value[0] >= after_column+1:
                new_value = value.copy()
                new_value[0] = m_colum
                after_end_names.update({key: new_value})
                m_colum += 1

    for nn_sheet in new_sheet_list:
        for key, val in after_end_names.items():
            nn_sheet.cell(row=first_row-1, column=val[0]).value = val[1]
            nn_sheet.column_dimensions[get_column_letter(
                val[0])].width = float(val[2])
            nn_sheet.row_dimensions[first_row-1].height = 130

    added_rows = []
    zz = first_row
    npp = 0
    double_sheet = new_sheet_list[0]
    for row, detail in details.items():
        if row in added_rows:
            pass
        else:
            added_rows += detail.doubles
            sum_details = 0
            npp += 1
            for nn_sheet in new_sheet_list:
                if nn_sheet.title == double_sheet_name:
                    zz = nn_sheet.max_row+1
                    for i in detail.doubles:
                        sum_details = sum_details + details.get(i).order
                        nn_sheet.cell(row=zz, column=after_end_names.get(
                            details.get(i).unit_row)[0]).value = details.get(i).order
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'npp')[0]).value = npp
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'types')[0]).value = detail.types
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'name')[0]).value = detail.name
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'knot')[0]).value = detail.knot
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'product')[0]).value = detail.product
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'order')[0]).value = sum_details
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'nzp')[0]).value = detail.nzp
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'massiv')[0]).value = detail.massiv
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'size_clean')[0]).value = detail.size_clean
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'size_workpiece')[0]).value = detail.size_workpiece
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'operations')[0]).value = detail.operations
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'shop')[0]).value = "-".join(detail.shop)
                    sum_details = 0
                if nn_sheet.title in detail.shop:
                    zz = nn_sheet.max_row+1
                    for i in detail.doubles:
                        sum_details = sum_details + details.get(i).order
                        nn_sheet.cell(row=zz, column=after_end_names.get(
                            details.get(i).unit_row)[0]).value = details.get(i).order
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'npp')[0]).value = npp
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'types')[0]).value = detail.types
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'name')[0]).value = detail.name
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'knot')[0]).value = detail.knot
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'product')[0]).value = detail.product
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'order')[0]).value = sum_details
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'nzp')[0]).value = detail.nzp
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'massiv')[0]).value = detail.massiv
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'size_clean')[0]).value = detail.size_clean
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'size_workpiece')[0]).value = detail.size_workpiece
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'operations')[0]).value = detail.operations
                    nn_sheet.cell(row=zz, column=after_end_names.get(
                        'shop')[0]).value = "-".join(detail.shop)
                    sum_details = 0

    for key, val in end_names.items():
        sheet.column_dimensions[get_column_letter(
            val[0])].width = float(val[2])

    sheet.row_dimensions[first_row-1].height = 120

    for nn_sheet in new_sheet_list:
        # Закрасил столбцы с узлами
        for colum in range(after_column+1, after_column+len(units)+1):
            for row in range(first_row, nn_sheet.max_row+1):
                nn_sheet.cell(row=row, column=colum).fill = fill_1

    for nn_sheet in new_sheet_list:
        for row in range(first_row, nn_sheet.max_row+1):
            nn_sheet.cell(row=row, column=after_end_names.get(
                'order')[0]).fill = fill_1

    for nn_sheet in new_sheet_list:
        for i in range(1, nn_sheet.max_column+1):
            nn_sheet.cell(row=first_row-1, column=i).fill = fill_1
            nn_sheet.cell(row=first_row-1, column=i).border = border_1
            nn_sheet.cell(row=first_row-1, column=i).font = font_1
            for key, value in after_end_names.items():
                if value[2] == 3.7:
                    nn_sheet.cell(row=first_row-1,
                                  column=value[0]).alignment = alignment_1
            for cell in range(first_row, nn_sheet.max_row+1):
                nn_sheet.cell(row=cell, column=i).border = border_1

    for nn_sheet in new_sheet_list:
        for coll in range(2, 4):
            for row in range(1, first_row-1):
                nn_sheet.cell(row=row, column=coll)._style = copy(
                    sheet.cell(row=row, column=coll)._style)
                nn_sheet.cell(row=row, column=coll).value = sheet.cell(
                    row=row, column=coll).value

    new_file_name = []
    for row in range(1, first_row-1):
        new_file_name.append(str(sheet.cell(row=row, column=3).value))

    folder_for_find_xlsx = os.path.dirname(
        os.path.abspath(inspect.stack()[0][1]))
    try:
        new_file_name[-1] = time.strftime('%Y.%m.%d',
                                          time.strptime(new_file_name[-1], '%Y-%m-%d %H:%M:%S'))
    except:
        pass

    new_file_name_path = folder_for_find_xlsx + "\\" + \
        exit_files_dir + '\\' + ' - '.join(new_file_name)+".xlsx"

    try:
        work_wb.save(new_file_name_path)
        logger.info("Записал файл: %s. Обработал за: %s секунд."
            % (
                new_file_name_path,
                round(time.process_time() - start_time, 3),
            )
            )
    except OSError as identifier:
        big_trabl(new_file_name_path, identifier)

def sum_files(work_dir, exit_files_dir):
    num_do_summ = 9
    num_posle_summ = 4
    header1 = []
    hed1 = []
    wb = openpyxl.Workbook()
    new_sheet = wb[wb.sheetnames[0]]  # берем первый лист в файле
    i = 0
    exit_mass = []
    for need_file in work_with_dir(work_dir):
        logger.info('Работаю с файлом: ' + work_dir + '\\' + need_file)
        work_wb = file_to_wb(work_dir + '\\' + need_file)
        sheet = work_wb[double_sheet_name]
        for row in sheet.iter_rows():
            o = []
            o.append(need_file)
            for z in row:
                o.append(z.value)
            if o[2] == "Служебная №:" or o[2] == "Заказчик:" or o[2] == "Заказ №:" or o[2] == "Отгрузка:":
                logger.info('Откидываю строку потому что: '+o[2])

            elif o[2] == "Наименование":
                for i in o[num_do_summ:len(o)-num_posle_summ]:
                    header1.append(str(i))

            else:
                exit_mass.append(o)

    header = ['Наименование', 'Позиция',
              'Кол-во узел (шт.)', 'Кол-во изделие (шт.)', 'Кол-во заказ (шт.)', 'Кол-во НЗП (шт.)', 'Массив']
    header2 = ['Размеры чист.', 'Размеры заготовки', 'Операции', 'Розцеховка']
    mass = [header + header1 + header2]

    max_elem = 0
    # iter_colls = 12 #количество столбцов без столбцов которые добавлены при поиске дублей
    for i in exit_mass:
        if max_elem < len(i):
            max_elem = len(i)
    # print(max_elem)
    files_list = []

    start = 0
    len_insert = 0
    append_mas = []
    exit_mass2 = []
    for i in exit_mass:
        if i[0] not in files_list:
            files_list.append(i[0])
            len_insert += len(i[num_do_summ:len(i)-num_posle_summ])
    for i in range(0, len_insert):
        append_mas.append(None)
    files_list.clear()
    files_list.append(exit_mass[0][0])
    start2 = 0
    for i in exit_mass:
        if i[0] not in files_list:
            files_list.append(i[0])
            start += start2
        else:
            start2 = len(i[num_do_summ:len(i)-num_posle_summ])
        sp1 = [None] * start
        sp2 = i[num_do_summ:len(i)-num_posle_summ]
        sp3 = [None] * (len_insert - len(sp2) - start)
        sp4 = i[0:num_do_summ]
        sp5 = i[0-num_posle_summ:]
        sp6 = sp4 + sp1 + sp2 + sp3 + sp5
        exit_mass2.append(sp6.copy())
        sp1.clear()
        sp2.clear()
        sp3.clear()
        sp4.clear()
        sp5.clear()
        sp6.clear()
    exit_mass.clear()
    hed1 = [None] * len(exit_mass2[0])
    for rowtbl in exit_mass2:
        # print(rowtbl)
        for i in range(num_do_summ, len(rowtbl)-num_posle_summ):
            if rowtbl[i] != None:
                hed1[i] = rowtbl[0]

    for row in [hed1[2:len(hed1)]]:
        new_sheet.append(row)
    for row in exit_mass2:
        # row = row[2:len(row)]
        del row[0]
        del row[0]
    for row in mass:
        new_sheet.append(row)
    for row in exit_mass2:
        new_sheet.append(row)

    # Расскрашиваем суммировочную таблицу
    logger.info('Расскрашиваем лист: %s' % ('Sheet'))
    for row in range(1, new_sheet.max_row+1):
        for col in range(1, new_sheet.max_column+1):
            new_sheet.cell(row=row, column=col).border = border_1

    for col in range(1, new_sheet.max_column+1):
        new_sheet.cell(row=1, column=col).fill = fill_1
        new_sheet.cell(row=1, column=col).font = font_1
        new_sheet.cell(row=2, column=col).fill = fill_1
        new_sheet.cell(row=2, column=col).font = font_1
        if col in [3, 4, 5, 6, 7]:
            new_sheet.cell(row=2, column=col).alignment = alignment_1
            new_sheet.column_dimensions[get_column_letter(col)].width = float(5)
        if col == 1:
            new_sheet.column_dimensions[get_column_letter(col)].width = float(20)
        if col == 2:
            new_sheet.column_dimensions[get_column_letter(col)].width = float(25)
        if col in [new_sheet.max_column, new_sheet.max_column-1, new_sheet.max_column-2, new_sheet.max_column-3]:
            new_sheet.column_dimensions[get_column_letter(col)].width = float(25)
        if col in list(range(8, new_sheet.max_column+1-4)):
            new_sheet.cell(row=1, column=col).comment = Comment(new_sheet.cell(row=1, column=col).value, "Find Doubles")
            new_sheet.cell(row=2, column=col).comment = Comment(new_sheet.cell(row=2, column=col).value, "Find Doubles")
            new_sheet.cell(row=2, column=col).alignment = alignment_1
            new_sheet.column_dimensions[get_column_letter(col)].width = float(5)
            new_sheet.cell(row=1, column=col).alignment = alignment_1
        new_sheet.row_dimensions[1].height = 60
        new_sheet.row_dimensions[2].height = 130

    try:
        wb.save(exit_files_dir + '\\' + exit_sum_file)
    except PermissionError as identifier:
        logger.critical('Нет доступа к файлу: %s, %s' % (exit_files_dir + '\\' + exit_sum_file, identifier))
        exit(0)

    work_wb = file_to_wb(exit_files_dir +'\\' + exit_sum_file)
    sheet = work_wb['Sheet']
    sum_mass = []
    def recell(itm):
        return itm.value
    for row in sheet.iter_rows():
        sum_mass.append(list(map(recell, row)))

    hhed1 = sum_mass[0]
    hhed2 = sum_mass[1]
    del sum_mass[0]
    del sum_mass[0]
    double_sheet = work_wb.create_sheet('Итог') # создаем новый лист в файле
    double_sheet.append(hhed1)
    double_sheet.append(hhed2)
    doubles_rows = []
    for row in sum_mass:
        exit_row = row
        if row[1] not in doubles_rows:
            ss = 0
            for row2 in sum_mass:
                if row[1] == row2[1]:
                    doubles_rows.append(row2[1])
                    ss += int(row2[4])
                    # сравниваем столбцы дублирующихся строк и выводим несовпадающие:
                    if row2[0] != exit_row[0]:
                        logger.warning('У %s %s не совпадает Наименование: "%s" и "%s"' %(row2[0], row2[1], row2[0], exit_row[0]))
                        exit_row[0] = '|'.join([str(row2[0]), str(exit_row[0]), 'W'])
                    # logger.debug(exit_row[-1])
                    if row2[-1] != exit_row[-1]:
                        logger.warning('У %s %s не совпадает Розцеховка: "%s" и "%s"' %(row2[0], row2[1], row2[-1], exit_row[-1]))
                        exit_row[-1] = '|'.join([str(row2[-1]), str(exit_row[-1]), 'W'])

                    if row2[-2] != exit_row[-2]:
                        logger.warning('У %s %s не совпадают Операции: "%s" и "%s"' %(row2[0], row2[1], row2[-2], exit_row[-2]))
                        if row2[-2] != None:
                            exit_row[-2] = '|'.join([str((row2[-2])), str(exit_row[-2]), 'W'])
                        # print(exit_row)

                    if row2[-3] != exit_row[-3]:
                        logger.warning('У %s %s не совпадают Размеры заготовки: "%s" и "%s"' %(row2[0], row2[1], row2[-3], exit_row[-3]))
                        exit_row[-3] = '|'.join([str(row2[-3]), str(exit_row[-3]), 'W'])

                    if row2[-4] != exit_row[-4]:
                        logger.warning('У %s %s не совпадают Размеры чист.: "%s" и "%s"' %(row2[0], row2[1], row2[-4], exit_row[-4]))
                        exit_row[-4] = '|'.join([str(row2[-4]), str(exit_row[-4]), 'W'])

                    for rr in range(0, len(row2)):
                        if row2[rr] != None and exit_row[rr] == None:
                            exit_row[rr] = row2[rr]

            exit_row[4] = ss
            ss = 0
            double_sheet.append(exit_row)

    try:
        work_wb.save(exit_files_dir + '\\' + exit_sum_file)
    except PermissionError as identifier:
        logger.critical('Нет доступа к файлу: %s' % (exit_files_dir + '\\' + exit_sum_file))
        exit(0)

    work_wb = file_to_wb(exit_files_dir +'\\' + exit_sum_file)
    sheet = work_wb['Итог']


    # Расскраска выходной таблицы
    logger.info('Расскрашиваем лист: %s' % ('Итог'))
    for row in range(1, sheet.max_row+1):
        for col in range(1, sheet.max_column+1):
            if col == 5:
                sheet.cell(row=row, column=col).fill = fill_1
            try:
                split_cell = sheet.cell(row=row, column=col).value.split('|')
                if len(split_cell) > 1:
                    sheet.cell(row=row, column=col).value = '|'.join(split_cell[:-1])
                    sheet.cell(row=row, column=col).fill = fill_2
            except:
                pass
            sheet.cell(row=row, column=col).border = border_1

    for col in range(1, sheet.max_column+1):
        sheet.cell(row=1, column=col).fill = fill_1
        sheet.cell(row=1, column=col).font = font_1
        sheet.cell(row=2, column=col).fill = fill_1
        sheet.cell(row=2, column=col).font = font_1
        if col in [3, 4, 5, 6, 7]:
            sheet.cell(row=2, column=col).alignment = alignment_1
            sheet.column_dimensions[get_column_letter(col)].width = float(5)
        if col == 1:
            sheet.column_dimensions[get_column_letter(col)].width = float(20)
        if col == 2:
            sheet.column_dimensions[get_column_letter(col)].width = float(25)
        if col in [sheet.max_column, sheet.max_column-1, sheet.max_column-2, sheet.max_column-3]:
            sheet.column_dimensions[get_column_letter(col)].width = float(25)
        if col in list(range(8, sheet.max_column+1-4)):
            sheet.cell(row=1, column=col).comment = Comment(sheet.cell(row=1, column=col).value, "Find Doubles")
            sheet.cell(row=2, column=col).comment = Comment(sheet.cell(row=2, column=col).value, "Find Doubles")
            sheet.cell(row=2, column=col).alignment = alignment_1
            sheet.column_dimensions[get_column_letter(col)].width = float(5)
            sheet.cell(row=1, column=col).alignment = alignment_1
        sheet.row_dimensions[1].height = 60
        sheet.row_dimensions[2].height = 130

    try:
        work_wb.save(exit_files_dir + '\\' + exit_sum_file)
    except PermissionError as identifier:
        logger.critical('Не получилось сохранить файл: %s' % (exit_files_dir + '\\' + exit_sum_file))
        exit(0)

if __name__ == '__main__':
    start_time = time.process_time()
    logger.info('Начинаю работу')
    try:
        # полный путь к папке из которой выполняется файл
        folder_for_find_xlsx = os.path.dirname(
            os.path.abspath(inspect.stack()[0][1]))

        # создаю папку "exit"
        mypath = folder_for_find_xlsx+'\\'+exit_files_dir
        if not os.path.isdir(mypath):
            logger.warning('Создаю папку "%s"'% (exit_files_dir))
            os.makedirs(mypath)

        # for xls_file in work_with_dir(folder_for_find_xlsx):
        #     print(xls_file)
        #     work(xls_file, folder_for_find_xlsx, exit_files_dir)

        try:
            for xls_file in work_with_dir(folder_for_find_xlsx):
                logger.info('Работаю с файлом: %s' % (xls_file))
                work(xls_file, folder_for_find_xlsx, exit_files_dir)
        except ValueError as trabl:
            big_trabl(xls_file, trabl)
            time.sleep(30)
        except FileNotFoundError as trabl:
            big_trabl(xls_file, trabl)
            time.sleep(30)
        except PermissionError as trabl:
            big_trabl(xls_file, trabl)
            time.sleep(30)

        folder_for_new_xlsx = folder_for_find_xlsx + '\\' + exit_files_dir

        try:
            sum_files(folder_for_new_xlsx, folder_for_find_xlsx)
        except KeyError as trabl:
            xls_file = 'сам знаешь'
            big_trabl(xls_file, trabl)

        # print('Можно закрывать.')
        logger.info('Можно закрывать. Работа закончена за: %s' % round(time.process_time() - start_time, 3))
        time.sleep(30)

    except BaseException as error:
        logger.critical('Ошибка: %s' % (error))
        time.sleep(30)