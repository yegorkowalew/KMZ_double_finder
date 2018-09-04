import openpyxl
from copy import copy
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

from datetime import datetime
import time
import os
import inspect
from modules.dirs.dir_work import user_input_work
# from modules.clear_operations import 

# def file_to_wb(file_name):
#     return openpyxl.load_workbook(filename = file_name)

# colontitul = [
#     "Служебная №:",
#     "Заказчик:",
#     "Заказ №:",
#     "Отгрузка:",
#     ]

# if __name__ == '__main__':
#     start_time = time.process_time() # Засекаем время
    
#     files = user_input_work()
#     work_wb = file_to_wb(files[0])
#     sheet = work_wb[work_wb.sheetnames[0]]
#     # print(sheet)
#     for row in range(1, 5):
#         for col in range(1, 10):
#             print(sheet.cell(row=row, column=col).value)

# for xls_file in files:
#     print(xls_file)
# Условия правильной работы скрипта:
#     Обрабатываемые поля:
#     Уз. - узел состящий из деталей
#         Столбец: "п/п" - Значение означающее что это узел "Уз."
#         Столбец: "Наименование" тип узла (указывать не обязательно)
#         Столбец: "Позиция" название узла (указывать обязательно)
#         Столбец: "Кол-во узел" (указывать не обязательно)
#         Столбец: "Кол-во изделие" (указывать не обязательно)
#         Столбец: "Кол-во заказ" (указывать не обязательно)
#         Столбец: "Кол-во НЗП" (указывать не обязательно)
#         Столбец: "Массив" (указывать не обязательно)
#         Столбец: "Размеры чист." Описание узла (указывать не обязательно, но желательно)
#         Столбец: "Размеры заготовки" (указывать не обязательно)
#         Столбец: "Операции" (указывать не обязательно)
#         Столбец: "Розцеховка" (указывать не обязательно)

#     Деталь - родитель Уз.
#         Столбец: "п/п" - Номер по порядку детали (Указывать не обязательно)
#         Столбец: "Наименование" тип узла (указывать обязательно)
#         Столбец: "Позиция" название детали (указывать обязательно)
#         Столбец: "Кол-во узел" (указывать не обязательно)
#         Столбец: "Кол-во изделие" (указывать не обязательно)
#         Столбец: "Кол-во заказ" по этому полю будут подсчитываться суммы дубликатов (указывать обязательно)
#         Столбец: "Кол-во НЗП" (указывать не обязательно)
#         Столбец: "Массив" (указывать не обязательно)
#         Столбец: "Размеры чист." (указывать не обязательно)
#         Столбец: "Размеры заготовки" (указывать не обязательно)
#         Столбец: "Операции" (указывать не обязательно)
#         Столбец: "Розцеховка" таблица будет разбита на листы в которых каждый цех по отдельности (указывать обязательно)

# === SETTINGS ===
parse_file = 'c:\\Users\\i.kovalenko\\parser\\find_doubles\\140.xlsx' # путь к файлу выгруженному с одноэса
first_row = 6
end_coll = 0
second_row = 0
after_column = 8
double_sheet_name = "Сортировка по узлам"
shop_list = []
new_sheet_list = []
end_names = {
        'npp': [1, "№ п/п", 3.7,],
        'types': [2, "Наименование", "21",],
        'name': [3, "Позиция", "36",],
        'knot': [4, "Кол-во узел (шт.)", 3.7,],
        'product': [5, "Кол-во изделие (шт.)", 3.7,],
        'order': [6, "Кол-во заказ (шт.)", 3.7,],
        'nzp': [7, "Кол-во НЗП (шт.)", 3.7,],
        'massiv': [8, "Массив", 3.7,],
        'size_clean': [9, "Размеры чист.", "26",],
        'size_workpiece': [10, "Размеры заготовки", "26",],
        'operations': [11, "Операции", "26",],
        'shop': [12, "Розцеховка", "12",],
        }

def if_num(cell):
    try:
        eval(cell)
        return True
    except:
        return False

class Unit:
    def __init__(self, types, name, knot, product, order, nzp, massiv, size_clean, size_workpiece, operations, shop, row):
        """Constructor for Unit"""
        self.types = types # Столбец "Наименование" (не обязательный)
        self.name = name # Столбец "Позиция" (обязательный)
        self.knot = knot # Столбец "Узел" (не обязательный)
        self.product = product # Столбец "Изделие" (не обязательный)
        self.order = order # Столбец "Заказ" (обязательный)
        self.nzp = nzp # Столбец: "Кол-во НЗП" (указывать не обязательно)
        self.massiv = massiv # Столбец: "Массив" (указывать не обязательно)
        self.size_clean = size_clean # Столбец: "Размеры чист." Описание узла (указывать не обязательно, но желательно)
        self.size_workpiece = size_workpiece # Столбец: "Размеры заготовки" (указывать не обязательно)
        self.operations = operations # Столбец: "Операции" (указывать не обязательно)
        self.shop = shop # Столбец: "Розцеховка" (указывать не обязательно)
        self.row = row # поле для хранения номера строки

class Detail:
    def __init__(self, npp, types, name, knot, product, order, nzp, massiv, size_clean, size_workpiece, operations, shop, doubles, row, unit_row):
        """Constructor for Detail"""
        self.npp = npp # "п/п" - Номер по порядку детали (Указывать не обязательно, желательно самому пронумеровать)
        self.types = types # Столбец: "Наименование" тип узла (указывать обязательно)
        self.name = name # Столбец "Позиция" (обязательный)
        self.knot = knot # Столбец "Узел" (не обязательный)
        self.product = product # Столбец "Изделие" (не обязательный)
        if if_num(order):
            self.order = eval(order) # Столбец: "Кол-во заказ" по этому полю будут подсчитываться суммы дубликатов (указывать обязательно)
        else:
            self.order = int(order)
        self.nzp = nzp # Столбец: "Кол-во НЗП" (указывать не обязательно)
        self.massiv = massiv # Столбец: "Массив" (указывать не обязательно)
        self.size_clean = size_clean # Столбец: "Размеры чист." (указывать не обязательно)
        self.size_workpiece = size_workpiece # Столбец: "Размеры заготовки" (указывать не обязательно)
        self.operations = operations # Столбец: "Операции" (указывать не обязательно)
        # self.shop Столбец: "Розцеховка" таблица будет разбита на листы в которых каждый цех по отдельности (указывать обязательно)
        self.shop = str(shop).split("-") # разбиваем строку вида 104-101, на список [104, 101]
        for st in self.shop:
            st = st.replace('/','')
        # print(self.shop)
        self.doubles = doubles # список дубликатов 
        self.row = row # поле для хранения номера строки
        self.unit_row = unit_row # Указываем строку родительского узла

def work_with_dir():
    folder_for_find_xlsx = os.path.dirname(os.path.abspath(inspect.stack()[0][1])) # полный путь к папке из которой выполняется файл
    list_files = [] # список файлов с полными путями
    for file in os.listdir(folder_for_find_xlsx): # во всех файлах папки
            if file.endswith(".xlsx"): # находим файл совпадающий с шаблоном
                list_files.append(os.path.join(folder_for_find_xlsx, file)) # добавляем файл, с полным путем, подходящий шаблону, в список
    return list_files

def file_to_wb(file_name):
    return openpyxl.load_workbook(filename = file_name)

if __name__ == '__main__':
    start_time = time.process_time() # Засекаем время
    xls_file = work_with_dir()[0] # первый файл в найденых файлах
    print('Работаю с файлом: ' + xls_file)

    work_wb = file_to_wb(xls_file) # из файла открываем книгу
    sheet = work_wb[work_wb.sheetnames[0]] # берем первый лист в файле
    if double_sheet_name in work_wb.sheetnames:
        double_sheet = work_wb[double_sheet_name]
        work_wb.remove(double_sheet)
        double_sheet = work_wb.create_sheet(double_sheet_name) # создаем новый лист в файле
        new_sheet_list.append(double_sheet)
    else:
        double_sheet = work_wb.create_sheet(double_sheet_name) # создаем новый лист в файле
        new_sheet_list.append(double_sheet)

    units = []
    details = {}

    for row in range(1, sheet.max_row): # нашел последнюю строку в таблице
        if sheet.cell(row=row, column=end_names.get('npp')[0]).value == "Уз.":
            first_row = row
            break
    
    for row in range(first_row, sheet.max_row): # нашел последнюю строку в таблице
        if sheet.cell(row=row, column=end_names.get('name')[0]).value:
            second_row = row+1

    print('Последняя обрабатываемая строка в таблице: '+str(second_row))

    for col in range(1, sheet.max_column): # нашел последний столбец в таблице
        end_coll = col

    for row in range(first_row, second_row):
        if sheet.cell(row=row, column=1).value == "Уз.":
            print('-')
            print('| Нашел узел: '+str(sheet.cell(row=row, column=3).value))
            unit_row = row
            npp = 0
            unit = Unit(
            sheet.cell(row=row, column=end_names.get('types')[0]).value,
            sheet.cell(row=row, column=end_names.get('name')[0]).value,
            sheet.cell(row=row, column=end_names.get('knot')[0]).value,
            sheet.cell(row=row, column=end_names.get('product')[0]).value,
            sheet.cell(row=row, column=end_names.get('order')[0]).value,
            sheet.cell(row=row, column=end_names.get('nzp')[0]).value,
            sheet.cell(row=row, column=end_names.get('massiv')[0]).value,
            sheet.cell(row=row, column=end_names.get('size_clean')[0]).value,
            sheet.cell(row=row, column=end_names.get('size_workpiece')[0]).value,
            sheet.cell(row=row, column=end_names.get('operations')[0]).value,
            sheet.cell(row=row, column=end_names.get('shop')[0]).value,
            row,)
            units.append(unit)
        else:
            print('-- Нашел деталь: '+sheet.cell(row=row, column=3).value)
            npp +=1
            unit = Detail(
            npp,
            sheet.cell(row=row, column=end_names.get('types')[0]).value,
            sheet.cell(row=row, column=end_names.get('name')[0]).value,
            sheet.cell(row=row, column=end_names.get('knot')[0]).value,
            sheet.cell(row=row, column=end_names.get('product')[0]).value,
            sheet.cell(row=row, column=end_names.get('order')[0]).value,
            sheet.cell(row=row, column=end_names.get('nzp')[0]).value,
            sheet.cell(row=row, column=end_names.get('massiv')[0]).value,
            sheet.cell(row=row, column=end_names.get('size_clean')[0]).value,
            sheet.cell(row=row, column=end_names.get('size_workpiece')[0]).value,
            sheet.cell(row=row, column=end_names.get('operations')[0]).value,
            sheet.cell(row=row, column=end_names.get('shop')[0]).value,
            [],
            row,
            unit_row)
            details.update({row:unit})

    for detail_for_add in details.values(): # нашел дубликаты и дописал их в поле .doubles
        for detail in details.values(): 
            if detail_for_add.name == detail.name:
                detail_for_add.doubles.append(detail.row)

    for detail in details.values(): # Список цехов
        for s in detail.shop:
            s = s.replace('/','')
            if s in shop_list:
                pass
            else:
                shop_list.append(s)
    # sheet = work_wb[work_wb.sheetnames[0]] # берем первый лист в файле
    # print(shop_list)
    for shop in shop_list:
        if shop in work_wb.sheetnames:
            double_sheet = work_wb[shop]
            work_wb.remove(double_sheet)
            double_sheet = work_wb.create_sheet(shop) # создаем новый лист в файле
            new_sheet_list.append(double_sheet)
        else:
            # print(shop)
            double_sheet = work_wb.create_sheet(shop) # создаем новый лист в файле
            new_sheet_list.append(double_sheet)

    after_end_names = {}
    colum = 0
    # Новый массив after_end_names со вставлеными столбцами уз.
    for key, value in end_names.items():
        if value[0] < after_column+1:
            after_end_names.update({key:value})
        else:
            if value[0] == after_column+1:
                colum = value[0]
                for i in units:
                    after_end_names.update({i.row:[colum, i.name, 3.7,]})
                    colum += 1
            if value[0] >= after_column+1:
                value[0] = colum
                after_end_names.update({key:value})
                colum += 1

    border_1 = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    fill_1 = PatternFill(fgColor='D9D9D9', fill_type = 'solid')

    font_1 = Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    alignment_1 = Alignment(text_rotation=180, 
                            wrap_text=False,
                            shrink_to_fit=False, 
                            indent=0
    )

    for nn_sheet in new_sheet_list:
        for key, val in after_end_names.items():
            nn_sheet.cell(row=first_row-1, column=val[0]).value = val[1]
            nn_sheet.column_dimensions[get_column_letter(val[0])].width = float(val[2])
            nn_sheet.row_dimensions[first_row-1].height = 120

    added_rows = []
    zz = first_row
    npp = 0
    double_sheet = new_sheet_list[0]
    for row, detail in details.items():
        if row in added_rows:
            pass
        else:
            added_rows += detail.doubles
            sum_details = 0
            npp += 1
            for nn_sheet in new_sheet_list:
                if nn_sheet.title == double_sheet_name:
                    zz = nn_sheet.max_row+1
                    for i in detail.doubles:
                        sum_details = sum_details + details.get(i).order
                        nn_sheet.cell(row=zz, column=after_end_names.get(details.get(i).unit_row)[0]).value = details.get(i).order
                    nn_sheet.cell(row=zz, column=after_end_names.get('npp')[0]).value = npp
                    nn_sheet.cell(row=zz, column=after_end_names.get('types')[0]).value = detail.types
                    nn_sheet.cell(row=zz, column=after_end_names.get('name')[0]).value = detail.name
                    nn_sheet.cell(row=zz, column=after_end_names.get('knot')[0]).value = detail.knot
                    nn_sheet.cell(row=zz, column=after_end_names.get('product')[0]).value = detail.product
                    nn_sheet.cell(row=zz, column=after_end_names.get('order')[0]).value = sum_details
                    nn_sheet.cell(row=zz, column=after_end_names.get('nzp')[0]).value = detail.nzp
                    nn_sheet.cell(row=zz, column=after_end_names.get('massiv')[0]).value = detail.massiv
                    nn_sheet.cell(row=zz, column=after_end_names.get('size_clean')[0]).value = detail.size_clean
                    nn_sheet.cell(row=zz, column=after_end_names.get('size_workpiece')[0]).value = detail.size_workpiece
                    nn_sheet.cell(row=zz, column=after_end_names.get('operations')[0]).value = detail.operations
                    nn_sheet.cell(row=zz, column=after_end_names.get('shop')[0]).value = "-".join(detail.shop)
                    sum_details = 0
                if nn_sheet.title in detail.shop:
                    zz = nn_sheet.max_row+1
                    for i in detail.doubles:
                        sum_details = sum_details + details.get(i).order
                        nn_sheet.cell(row=zz, column=after_end_names.get(details.get(i).unit_row)[0]).value = details.get(i).order
                    nn_sheet.cell(row=zz, column=after_end_names.get('npp')[0]).value = npp
                    nn_sheet.cell(row=zz, column=after_end_names.get('types')[0]).value = detail.types
                    nn_sheet.cell(row=zz, column=after_end_names.get('name')[0]).value = detail.name
                    nn_sheet.cell(row=zz, column=after_end_names.get('knot')[0]).value = detail.knot
                    nn_sheet.cell(row=zz, column=after_end_names.get('product')[0]).value = detail.product
                    nn_sheet.cell(row=zz, column=after_end_names.get('order')[0]).value = sum_details
                    nn_sheet.cell(row=zz, column=after_end_names.get('nzp')[0]).value = detail.nzp
                    nn_sheet.cell(row=zz, column=after_end_names.get('massiv')[0]).value = detail.massiv
                    nn_sheet.cell(row=zz, column=after_end_names.get('size_clean')[0]).value = detail.size_clean
                    nn_sheet.cell(row=zz, column=after_end_names.get('size_workpiece')[0]).value = detail.size_workpiece
                    nn_sheet.cell(row=zz, column=after_end_names.get('operations')[0]).value = detail.operations
                    nn_sheet.cell(row=zz, column=after_end_names.get('shop')[0]).value = "-".join(detail.shop)
                    sum_details = 0

    for key, val in end_names.items():
        sheet.column_dimensions[get_column_letter(val[0])].width = float(val[2])

    sheet.row_dimensions[first_row-1].height = 120
    
    for nn_sheet in new_sheet_list:
        for colum in range(after_column+1, after_column+len(units)+1): # Закрасил столбцы с узлами
            for row in range(first_row, nn_sheet.max_row+1):
                nn_sheet.cell(row=row, column=colum).fill = fill_1
    
    for nn_sheet in new_sheet_list:
        for row in range(first_row, nn_sheet.max_row+1):
            nn_sheet.cell(row=row, column=after_end_names.get('order')[0]).fill = fill_1

    for nn_sheet in new_sheet_list:
        for i in range(1, nn_sheet.max_column+1):
            nn_sheet.cell(row=first_row-1, column=i).fill = fill_1
            nn_sheet.cell(row=first_row-1, column=i).border = border_1
            nn_sheet.cell(row=first_row-1, column=i).font = font_1
            for key, value in after_end_names.items():
                if value[2] == 3.7:
                    nn_sheet.cell(row=first_row-1, column=value[0]).alignment = alignment_1
            for cell in range(first_row, nn_sheet.max_row+1):
                nn_sheet.cell(row=cell, column=i).border = border_1

    for nn_sheet in new_sheet_list:
        for coll in range(2, 4):
            for row in range(1, first_row-1):
                nn_sheet.cell(row=row, column=coll)._style = copy(sheet.cell(row=row, column=coll)._style)
                nn_sheet.cell(row=row, column=coll).value = sheet.cell(row=row, column=coll).value

    new_file_name = []
    for row in range(1, first_row-1):
                new_file_name.append(str(sheet.cell(row=row, column=3).value))

    folder_for_find_xlsx = os.path.dirname(os.path.abspath(inspect.stack()[0][1]))
    try:
        new_file_name[-1] = time.strftime('%Y.%m.%d', time.strptime(new_file_name[-1], '%Y-%m-%d %H:%M:%S'))
    except:
        pass

    new_file_name_path = folder_for_find_xlsx + "\\" + ' - '.join(new_file_name)+".xlsx"

    work_wb.save(new_file_name_path)
    print('Сохраняю')
    print("Сохранил файл с новым именем: %s. \nЗа %s секунд. Время: %s"
        %(
            new_file_name_path,
            round(time.process_time() - start_time, 3),
            datetime.strftime(datetime.now(), "%H:%M:%S"),
            )
        )
    print('Можно закрывать окно')
    time.sleep(30)